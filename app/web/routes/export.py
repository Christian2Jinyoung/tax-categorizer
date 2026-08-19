import io
from datetime import datetime
from typing import Optional
from urllib.parse import quote

import openpyxl
import pandas as pd
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import select

from app.config import PROJECT_ROOT
from app.db import get_session
from app.models import LineItem, SourceFile, UploadBatch

# review_status values that mean a human has already made the deductible call - either
# by clicking through /review or by importing a checked-off export (see import_checks
# below) - as opposed to "auto_ok"/"needs_review", which are just the categorizer's guess.
REVIEWED_STATUSES = ("reviewed_confirmed", "reviewed_overridden")

router = APIRouter()

# Per-column Excel number formats, applied wherever a sheet has one of these headers.
COLUMN_NUMBER_FORMATS = {
    "Date": "mm/dd/yyyy",
    "Price": '"$"#,##0.00',
    "Deduction %": "0%",
}
MAX_COLUMN_WIDTH = 60

# Excel's built-in "Good"/"Bad" cell-style colors, for the manual Deductible column.
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str) -> None:
    """Writes df to the workbook, then widens every column to fit its content and
    applies currency/date/percent formatting to columns named in COLUMN_NUMBER_FORMATS.

    If the sheet has a "Deductible" column, also adds live conditional formatting so
    typing "y" or "x" into that column colors the whole row green/red in Excel itself -
    no re-export needed to see the color change.
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]

    for col_idx, column_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        cell_lengths = [len(str(v)) for v in df[column_name]] if len(df) else []
        width = max([len(str(column_name)), *cell_lengths]) + 2
        worksheet.column_dimensions[col_letter].width = min(width, MAX_COLUMN_WIDTH)

        number_format = COLUMN_NUMBER_FORMATS.get(column_name)
        if number_format:
            for row_idx in range(2, len(df) + 2):  # row 1 is the header
                worksheet.cell(row=row_idx, column=col_idx).number_format = number_format

    if "Deductible" in df.columns and len(df):
        deductible_col_letter = get_column_letter(df.columns.get_loc("Deductible") + 1)
        last_col_letter = get_column_letter(len(df.columns))
        cell_range = f"A2:{last_col_letter}{len(df) + 1}"
        worksheet.conditional_formatting.add(
            cell_range, FormulaRule(formula=[f'LOWER(${deductible_col_letter}2)="y"'], fill=GREEN_FILL)
        )
        worksheet.conditional_formatting.add(
            cell_range, FormulaRule(formula=[f'LOWER(${deductible_col_letter}2)="x"'], fill=RED_FILL)
        )

# Every export is also written here as a standing copy on disk, in addition to the
# browser download - contains real purchase data, so it's gitignored (see .gitignore).
EXPORT_DIR = PROJECT_ROOT / "app" / "export"


def _save_export_copy(buffer: io.BytesIO, filename: str) -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    (EXPORT_DIR / filename).write_bytes(buffer.getvalue())


def _source_filenames(session, items: list[LineItem]) -> dict[int, str]:
    source_file_ids = {item.source_file_id for item in items}
    if not source_file_ids:
        return {}
    rows = session.exec(
        select(SourceFile.id, SourceFile.original_filename).where(SourceFile.id.in_(source_file_ids))
    ).all()
    return dict(rows)


@router.get("/export/{batch_id}.csv")
def export_csv(batch_id: int):
    with get_session() as session:
        batch = session.get(UploadBatch, batch_id)
        if batch is None:
            raise HTTPException(status_code=404, detail="Batch not found")
        items = session.exec(
            select(LineItem).where(LineItem.batch_id == batch_id).order_by(LineItem.id)
        ).all()

    rows = [
        {
            "Date": item.date,
            "Vendor": item.vendor,
            "Description": item.description,
            "Amount": item.amount,
            "Category": item.category_label,
            "Deductible": item.deductible,
            "Deduction %": item.deduction_pct,
            "Deductible Amount": (item.amount or 0) * (item.deduction_pct or 0) if item.deductible else 0,
            "Method": item.categorization_method,
            "Confidence": item.categorization_confidence,
            "Rationale": item.rationale,
            "Review Status": item.review_status,
        }
        for item in items
    ]
    df = pd.DataFrame(rows)
    buffer = io.StringIO()
    df.to_csv(buffer, index=False)

    filename = f"tax_categorizer_batch_{batch_id}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# The primary sheet for manual review: every real line item, regardless of what (if
# anything) the categorizer decided about deductibility - the point is the user checks
# each one themselves rather than trusting an auto-generated Deductible/Not-Deductible split.
# "Deductible" is the leftmost column - the user fills in "y" (deductible) or "x" (not
# deductible) per row, which live-colors the row via conditional formatting (see
# _write_sheet). It's blank only for rows nobody has reviewed yet; a row already
# reviewed (via /review or a prior round-trip through import_checks below) is
# pre-filled with its stored y/x so re-exporting never erases past review work.
ALL_ITEMS_COLUMNS = [
    "Deductible", "Item UID", "Date", "Full Name", "Abbreviated Name (Receipt)", "Quantity", "Price",
    "Source File",
]


def _all_items_row(item: LineItem, source_filenames: dict[int, str]) -> dict:
    reviewed = item.review_status in REVIEWED_STATUSES
    return {
        "Deductible": ("y" if item.deductible else "x") if reviewed else None,
        "Item UID": item.item_uid,
        "Date": item.date,
        "Full Name": item.description,
        "Abbreviated Name (Receipt)": item.abbreviated_description,
        "Quantity": item.quantity or 1,
        "Price": item.amount,
        "Source File": source_filenames.get(item.source_file_id),
    }


def _all_items_sheet(items: list[LineItem], source_filenames: dict[int, str]) -> pd.DataFrame:
    if not items:
        return pd.DataFrame(columns=ALL_ITEMS_COLUMNS)
    return pd.DataFrame([_all_items_row(item, source_filenames) for item in items], columns=ALL_ITEMS_COLUMNS)


SPREADSHEET_COLUMNS = [
    "Item UID", "Date", "Full Name", "Abbreviated Name (Receipt)", "Quantity", "Price", "Vendor",
    "Category", "Deduction %", "Needs Review", "Method", "Rationale", "Source File",
]


def _spreadsheet_row(item: LineItem, source_filenames: dict[int, str]) -> dict:
    return {
        "Item UID": item.item_uid,
        "Date": item.date,
        "Full Name": item.description,
        "Abbreviated Name (Receipt)": item.abbreviated_description,
        "Quantity": item.quantity or 1,
        "Price": item.amount,
        "Vendor": item.vendor,
        "Category": item.category_label,
        "Deduction %": item.deduction_pct,
        "Needs Review": item.review_status == "needs_review",
        "Method": item.categorization_method,
        "Rationale": item.rationale,
        "Source File": source_filenames.get(item.source_file_id),
    }


def _spreadsheet_sheet(items: list[LineItem], source_filenames: dict[int, str]) -> pd.DataFrame:
    if not items:
        return pd.DataFrame(columns=SPREADSHEET_COLUMNS)
    return pd.DataFrame([_spreadsheet_row(item, source_filenames) for item in items], columns=SPREADSHEET_COLUMNS)


def _build_workbook(items: list[LineItem], source_filenames: dict[int, str]) -> io.BytesIO:
    """The final spreadsheet. Leads with "All Items" - every real line item with its full
    resolved name, its original receipt abbreviation, quantity, price, date, and source
    receipt file, for manual review; the categorizer's guess is never pre-filled here for
    a row that hasn't been reviewed yet, only carried over for one that already has (see
    _all_items_row). The categorizer's own Tax Deductible / Not Deductible / Uncategorized
    split follows for reference.
    """
    all_items = [item for item in items if not (item.vendor is None and item.amount is None)]
    deductible = [item for item in items if item.deductible is True]
    non_deductible = [item for item in items if item.deductible is False]
    uncategorized = [item for item in items if item.deductible is None]
    deductible_total = sum((item.amount or 0) * (item.deduction_pct or 0) for item in deductible)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(writer, _all_items_sheet(all_items, source_filenames), "All Items")
        _write_sheet(
            writer,
            pd.DataFrame(
                [
                    {"Metric": "Total line items", "Value": len(all_items)},
                    {"Metric": "Deductible line items (per agent)", "Value": len(deductible)},
                    {"Metric": "Non-deductible line items (per agent)", "Value": len(non_deductible)},
                    {"Metric": "Uncategorized / extraction issues", "Value": len(uncategorized)},
                    {"Metric": "Estimated deductible total (per agent)", "Value": round(deductible_total, 2)},
                ]
            ),
            "Summary",
        )
        _write_sheet(writer, _spreadsheet_sheet(deductible, source_filenames), "Tax Deductible")
        _write_sheet(writer, _spreadsheet_sheet(non_deductible, source_filenames), "Not Deductible")
        if uncategorized:
            _write_sheet(writer, _spreadsheet_sheet(uncategorized, source_filenames), "Uncategorized")
    buffer.seek(0)
    return buffer


@router.get("/export/all.xlsx")
def export_all_xlsx():
    """Spans every batch ever ingested, not just one run. Ingestion already dedups by
    filename (a receipt already processed in an earlier batch is skipped, not reprocessed),
    so every real item exists exactly once in the database - this just stops the export
    from being limited to whichever single batch you happen to be looking at.

    Registered before the /export/{batch_id}.xlsx route below: FastAPI matches path
    routes in registration order, and "all" isn't a valid int batch_id, so this specific
    route must come first or every request here 422s trying to parse "all" as an int.
    """
    with get_session() as session:
        items = session.exec(select(LineItem).order_by(LineItem.date, LineItem.item_uid)).all()
        source_filenames = _source_filenames(session, items)

    buffer = _build_workbook(items, source_filenames)
    _save_export_copy(buffer, "tax_categorizer_all_items.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tax_categorizer_all_items.xlsx"'},
    )


@router.get("/export/{batch_id}.xlsx")
def export_xlsx(batch_id: int):
    with get_session() as session:
        batch = session.get(UploadBatch, batch_id)
        if batch is None:
            raise HTTPException(status_code=404, detail="Batch not found")
        items = session.exec(
            select(LineItem).where(LineItem.batch_id == batch_id).order_by(LineItem.id)
        ).all()
        source_filenames = _source_filenames(session, items)

    buffer = _build_workbook(items, source_filenames)
    filename = f"tax_categorizer_batch_{batch_id}.xlsx"
    _save_export_copy(buffer, filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _normalize_deductible_mark(value: object) -> Optional[bool]:
    text = str(value).strip().lower() if value is not None else ""
    if text in ("y", "yes"):
        return True
    if text in ("x", "n", "no"):
        return False
    return None  # blank, or an unrecognized value - treated as "not reviewed"


@router.post("/export/import-checks")
async def import_checks(file: UploadFile = File(...)):
    """Reads back a previously exported "All Items" sheet that's had its Deductible
    column hand-filled in with y/x, and persists those calls to the database (matched
    by Item UID) so they survive future re-exports instead of only ever living in that
    one Excel file - see _all_items_row, which now pre-fills already-reviewed rows
    instead of always blanking the column.
    """
    contents = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user, not a crash
        return RedirectResponse(url=f"/?error={quote(f'Could not read {file.filename}: {exc}')}", status_code=303)

    if "All Items" not in workbook.sheetnames:
        return RedirectResponse(
            url=f"/?error={quote(f'{file.filename} has no \"All Items\" sheet')}", status_code=303
        )

    sheet = workbook["All Items"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if "Item UID" not in header or "Deductible" not in header:
        return RedirectResponse(
            url=f"/?error={quote('All Items sheet is missing the Item UID or Deductible column')}",
            status_code=303,
        )
    uid_col = header.index("Item UID")
    deductible_col = header.index("Deductible")

    marks: dict[str, bool] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        uid = row[uid_col] if uid_col < len(row) else None
        mark = _normalize_deductible_mark(row[deductible_col] if deductible_col < len(row) else None)
        if uid and mark is not None:
            marks[str(uid)] = mark

    updated = 0
    not_found = 0
    with get_session() as session:
        for uid, is_deductible in marks.items():
            item = session.exec(select(LineItem).where(LineItem.item_uid == uid)).first()
            if item is None:
                not_found += 1
                continue
            item.deductible = is_deductible
            if is_deductible and not item.deduction_pct:
                item.deduction_pct = 1.0
            item.categorization_method = item.categorization_method or "manual_override"
            item.review_status = "reviewed_overridden"
            item.reviewed_at = datetime.utcnow()
            item.updated_at = datetime.utcnow()
            session.add(item)
            updated += 1
        session.commit()

    message = f"Imported {updated} checked row(s) from {file.filename}."
    if not_found:
        message += f" {not_found} Item UID(s) from the file weren't found (skipped)."
    return RedirectResponse(url=f"/?message={quote(message)}", status_code=303)
